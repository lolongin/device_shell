"""Safe preview parsing for spreadsheet-backed device inventory imports."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from device_tui.domain.devices.models import Device


MAX_IMPORT_BYTES = 20 * 1024 * 1024
MAX_IMPORT_ROWS = 20_000
MAX_IMPORT_COLUMNS = 100
PREVIEW_ROW_LIMIT = 15


class DeviceImportError(Exception):
    """Raised when an import file cannot produce a safe device preview."""


@dataclass(frozen=True, slots=True)
class DeviceImportIssue:
    row: int
    message: str


@dataclass(frozen=True, slots=True)
class ParsedDeviceImport:
    source_path: Path
    sheet_name: str
    headers: tuple[str, ...]
    devices: tuple[Device, ...]
    total_rows: int
    skipped_rows: int
    errors: tuple[DeviceImportIssue, ...]
    warnings: tuple[str, ...]
    preview_rows: tuple[dict[str, str], ...]


_ALIASES: dict[str, set[str]] = {
    "id": {"id", "deviceid", "设备id", "设备编号", "设备编码", "编号"},
    "name": {"name", "devicename", "displayname", "设备名称", "设备名", "名称"},
    "domain": {"domain", "domainname", "领域", "域", "业务域", "分组"},
    "device_type": {"type", "kind", "devicetype", "设备类型", "类型"},
    "cpu": {"cpu", "cpuarch", "架构", "cpu架构"},
    "status": {"status", "statuslabel", "状态"},
    "owner": {"owner", "user", "占用人", "负责人"},
    "host": {"ip", "host", "deviceip", "managementip", "设备ip", "管理ip", "主机"},
    "ssh_ip": {"sship", "sshhost", "ssh地址", "sship地址"},
    "ssh_port": {"sshport", "ssh端口"},
    "ssh_username": {"sshusername", "sshuser", "ssh账号", "ssh用户"},
    "telnet_ip": {"telnetip", "telnethost", "telnet地址"},
    "telnet_port": {"telnetport", "telnet端口"},
    "username": {"username", "user", "账号", "用户名", "登录账号"},
    "serial_ip": {"serialip", "serialhost", "串口地址", "串口服务器"},
    "serial_port": {"serialport", "串口端口"},
    "serial_username": {"serialusername", "serialuser", "串口账号", "串口用户"},
    "vendor": {"vendor", "厂商", "厂家"},
    "model": {"model", "型号", "设备型号", "板类型"},
    "site": {"site", "站点", "位置"},
    "rack": {"rack", "机架", "机柜", "slot", "槽位"},
    "version": {"version", "版本", "软件版本"},
    "notes": {"notes", "remark", "remarks", "备注", "说明"},
    "board_id": {"boardid", "board", "单板id", "板卡id", "板编号"},
}
_PASSWORD_HEADERS = {
    "password", "passwd", "pwd", "密码", "登录密码", "sshpassword", "telnetpassword",
    "serialpassword", "ssh密码", "telnet密码", "串口密码",
}


def parse_device_import(path: Path) -> ParsedDeviceImport:
    resolved = path.expanduser().resolve()
    if not resolved.is_file():
        raise DeviceImportError("导入文件不存在。")
    size = resolved.stat().st_size
    if size <= 0:
        raise DeviceImportError("导入文件为空。")
    if size > MAX_IMPORT_BYTES:
        raise DeviceImportError("导入文件不能超过 20 MB。")
    suffix = resolved.suffix.casefold()
    if suffix == ".xlsx":
        sheet_name, rows = _xlsx_rows(resolved)
    elif suffix in {".csv", ".tsv"}:
        sheet_name, rows = _delimited_rows(resolved, suffix)
    else:
        raise DeviceImportError("仅支持 .xlsx、.csv 和 .tsv 文件；旧版 .xls 请先另存为 .xlsx。")
    return _parse_rows(resolved, sheet_name, rows)


def _xlsx_rows(path: Path) -> tuple[str, Iterable[tuple[object, ...]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - packaging contract
        raise DeviceImportError("Excel 读取组件不可用。") from exc
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise DeviceImportError("无法读取 Excel 文件，请确认文件未损坏且未加密。") from exc
    sheet = next((item for item in workbook.worksheets if item.sheet_state == "visible"), None)
    if sheet is None:
        workbook.close()
        raise DeviceImportError("Excel 中没有可见工作表。")

    def rows() -> Iterable[tuple[object, ...]]:
        try:
            yield from sheet.iter_rows(values_only=True)
        finally:
            workbook.close()

    return sheet.title, rows()


def _delimited_rows(path: Path, suffix: str) -> tuple[str, Iterable[tuple[str, ...]]]:
    raw = path.read_bytes()
    text = ""
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if not text:
        raise DeviceImportError("无法识别 CSV/TSV 文件编码，请使用 UTF-8 或 GB18030。")
    delimiter = "\t" if suffix == ".tsv" else ","
    if suffix == ".csv":
        try:
            delimiter = csv.Sniffer().sniff(text[:8192], delimiters=",;\t").delimiter
        except csv.Error:
            delimiter = ","
    return path.stem, (
        tuple(row) for row in csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)
    )


def _parse_rows(
    path: Path,
    sheet_name: str,
    source_rows: Iterable[tuple[object, ...]],
) -> ParsedDeviceImport:
    iterator = iter(source_rows)
    header_row: tuple[object, ...] | None = None
    header_number = 0
    for row_number, row in enumerate(iterator, start=1):
        if any(_cell_text(value) for value in row):
            header_row = row
            header_number = row_number
            break
    if header_row is None:
        raise DeviceImportError("文件中没有可识别的表头。")
    if len(header_row) > MAX_IMPORT_COLUMNS:
        raise DeviceImportError("导入表不能超过 100 列。")
    headers = tuple(_cell_text(value) or f"未命名列{index + 1}" for index, value in enumerate(header_row))
    mapping = _column_mapping(headers)
    if "id" not in mapping and "name" not in mapping:
        raise DeviceImportError("表格至少需要“设备编号/ID”或“设备名称”列。")
    password_columns = [
        header for header in headers if _normalized_header(header) in _PASSWORD_HEADERS
    ]
    warnings: list[str] = []
    if password_columns:
        warnings.append("检测到密码列，已按安全策略忽略，不会写入本地数据库。")
    devices: list[Device] = []
    errors: list[DeviceImportIssue] = []
    preview_rows: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    total_rows = 0
    skipped_rows = 0
    for row_number, row in enumerate(iterator, start=header_number + 1):
        if not any(_cell_text(value) for value in row):
            continue
        total_rows += 1
        if total_rows > MAX_IMPORT_ROWS:
            raise DeviceImportError("单次导入不能超过 20,000 行。")
        try:
            values = {
                field: _cell_text(row[index]) if index < len(row) else ""
                for field, index in mapping.items()
            }
            device = _device_from_values(values)
            key = (device.id.casefold(), device.board_id.casefold())
            if key in seen:
                raise ValueError(f"设备 ID 与单板 ID 重复：{device.id} / {device.board_id or '-'}")
            seen.add(key)
            devices.append(device)
            if len(preview_rows) < PREVIEW_ROW_LIMIT:
                preview_rows.append({
                    "id": device.id,
                    "name": device.name,
                    "domain": device.domain,
                    "device_type": device.device_type,
                    "host": device.ssh_ip or device.telnet_ip or device.serial_ip,
                    "status": device.status,
                })
        except ValueError as exc:
            skipped_rows += 1
            errors.append(DeviceImportIssue(row=row_number, message=str(exc)))
    if not devices:
        detail = errors[0].message if errors else "没有数据行"
        raise DeviceImportError(f"没有可导入的有效设备：{detail}")
    return ParsedDeviceImport(
        source_path=path,
        sheet_name=sheet_name,
        headers=headers,
        devices=tuple(devices),
        total_rows=total_rows,
        skipped_rows=skipped_rows,
        errors=tuple(errors),
        warnings=tuple(warnings),
        preview_rows=tuple(preview_rows),
    )


def _column_mapping(headers: tuple[str, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = _normalized_header(header)
        for field, aliases in _ALIASES.items():
            if field not in mapping and normalized in aliases:
                mapping[field] = index
                break
    return mapping


def _device_from_values(values: dict[str, str]) -> Device:
    device_id = values.get("id", "").strip()
    name = values.get("name", "").strip()
    if not device_id and not name:
        raise ValueError("设备编号和设备名称不能同时为空。")
    device_id = device_id or name
    name = name or device_id
    if len(device_id) > 160:
        raise ValueError("设备编号不能超过 160 个字符。")
    generic_host = values.get("host", "").strip()
    ssh_ip = values.get("ssh_ip", "").strip() or generic_host
    telnet_ip = values.get("telnet_ip", "").strip() or generic_host
    serial_ip = values.get("serial_ip", "").strip()
    return Device(
        id=device_id,
        name=name,
        domain=values.get("domain", "").strip() or "导入设备",
        device_type=values.get("device_type", "").strip(),
        cpu=values.get("cpu", "").strip(),
        status=values.get("status", "").strip() or "空闲",
        owner=values.get("owner", "").strip() or None,
        ssh_ip=ssh_ip,
        telnet_ip=telnet_ip,
        username=values.get("username", "").strip(),
        password="",
        vendor=values.get("vendor", "").strip(),
        model=values.get("model", "").strip(),
        site=values.get("site", "").strip(),
        rack=values.get("rack", "").strip(),
        version=values.get("version", "").strip(),
        notes=values.get("notes", "").strip(),
        board_id=values.get("board_id", "").strip(),
        ssh_port=_port(values.get("ssh_port", ""), 22, "SSH"),
        telnet_port=_port(values.get("telnet_port", ""), 23, "Telnet"),
        ssh_username=values.get("ssh_username", "").strip() or values.get("username", "").strip(),
        ssh_password="",
        serial_ip=serial_ip,
        serial_port=_port(values.get("serial_port", ""), 23, "串口"),
        serial_username=values.get("serial_username", "").strip(),
        serial_password="",
        supports_power_off=False,
        extra={"supports_occupancy": False, "imported": True},
    )


def _port(value: str, default: int, label: str) -> int:
    if not value:
        return default
    try:
        parsed = int(float(value))
    except ValueError as exc:
        raise ValueError(f"{label} 端口不是有效数字：{value}") from exc
    if parsed < 1 or parsed > 65535:
        raise ValueError(f"{label} 端口必须在 1-65535 之间。")
    return parsed


def _normalized_header(value: object) -> str:
    return re.sub(r"[\s_\-./\\（）()]+", "", _cell_text(value).casefold())


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
